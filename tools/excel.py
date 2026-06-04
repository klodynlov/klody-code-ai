"""GÃĐnÃĐration de classeurs Excel (.xlsx) tÃĐlÃĐchargeables.

Le fichier est ÃĐcrit dans `config.DOWNLOADS_DIR` (servi par l'API sur
`/api/files/<nom>`) â JAMAIS sous le projet ni une racine arbitraire : le nom
est rÃĐduit Ã  un basename assaini et l'extension forcÃĐe Ã  `.xlsx`, donc aucune
traversÃĐe de chemin n'est possible.

`openpyxl` est une dÃĐpendance optionnelle au sens du code (mÃŠme pattern que
`tools/audio.py` avec librosa) : son absence renvoie une erreur lisible plutÃīt
que de planter l'agent. Elle est nÃĐanmoins dÃĐclarÃĐe dans requirements.txt car
la gÃĐnÃĐration Excel est une fonctionnalitÃĐ de premier plan.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from config import DOWNLOADS_DIR

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - dÃĐpend de l'environnement
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# Garde-fous (mÃĐmoire / temps de gÃĐnÃĐration).
_MAX_ROWS = 100_000
_MAX_COLS = 1_000
_MAX_COL_WIDTH = 60

# CaractÃĻres interdits dans un nom de fichier â on n'en garde qu'un sous-ensemble
# sÃŧr. `\w` (mode Unicode par dÃĐfaut) garde les lettres accentuÃĐes : appli FR, on
# ne veut pas mutiler ÂŦ rÃĐsumÃĐ.xlsx Âŧ en ÂŦ r_sum_.xlsx Âŧ. La protection
# anti-traversÃĐe ne repose PAS sur cette regex (cosmÃĐtique) mais sur
# `Path(...).name` + `resolve()` + vÃĐrif du dossier parent dans `generate_excel`.
_UNSAFE_NAME = re.compile(r"[^\w.\- ]+")
# CaractÃĻres interdits par Excel dans un titre d'onglet.
_UNSAFE_SHEET = re.compile(r"[\[\]:*?/\\]")


def _safe_filename(filename: str) -> str:
    """RÃĐduit `filename` Ã  un basename sÃŧr terminÃĐ par `.xlsx`.

    Vire tout composant de chemin (anti-traversÃĐe), nettoie les caractÃĻres
    exotiques et force l'extension `.xlsx`.
    """
    base = Path(str(filename or "").strip()).name
    base = _UNSAFE_NAME.sub("_", base).strip(". ")
    stem = Path(base).stem or "export"
    return f"{stem}.xlsx"


def _safe_sheet_title(name: Any, used: set[str]) -> str:
    """Titre d'onglet valide Excel (âĪ 31 car., sans `[]:*?/\\`), unique dans `used`."""
    title = _UNSAFE_SHEET.sub(" ", str(name or "Feuille")).strip()[:31] or "Feuille"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        n += 1
    used.add(title.lower())
    return title


def _coerce(value: Any) -> Any:
    """Valeur ÃĐcrivable telle quelle par openpyxl ; le reste est stringifiÃĐ."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _normalize_sheets(sheets: Any) -> list[dict]:
    """Normalise l'entrÃĐe en `[{name, columns, rows}]`.

    TolÃĻre : une feuille unique (dict), une liste de feuilles, et des lignes
    fournies en liste de dicts (les clÃĐs deviennent les en-tÃŠtes).
    """
    if sheets is None:
        return []
    if isinstance(sheets, dict):
        sheets = [sheets]
    if not isinstance(sheets, list):
        return []

    out: list[dict] = []
    for i, sh in enumerate(sheets):
        if not isinstance(sh, dict):
            continue
        name = sh.get("name") or f"Feuille{i + 1}"
        columns = sh.get("columns")
        rows = sh.get("rows") or []
        if not isinstance(rows, list):
            rows = []
        # Lignes en liste de dicts â en-tÃŠtes dÃĐrivÃĐs de l'union ordonnÃĐe des clÃĐs.
        if rows and isinstance(rows[0], dict):
            if not columns:
                seen: dict[str, None] = {}
                for r in rows:
                    if isinstance(r, dict):
                        for k in r:
                            seen.setdefault(str(k), None)
                columns = list(seen)
            rows = [[r.get(c) for c in columns] for r in rows if isinstance(r, dict)]
        out.append({"name": name, "columns": columns, "rows": rows})
    return out


def _autosize(ws: Any, columns: Any, rows: list) -> None:
    """Largeur de colonne â longueur du contenu le plus large (bornÃĐe)."""
    widths: dict[int, int] = {}

    def consider(idx: int, value: Any) -> None:
        length = len(str(value)) if value is not None else 0
        if length > widths.get(idx, 0):
            widths[idx] = length

    if columns:
        for i, c in enumerate(columns):
            consider(i, c)
    for r in rows:
        cells = r if isinstance(r, (list, tuple)) else [r]
        for i, c in enumerate(cells):
            consider(i, c)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w + 2, 8), _MAX_COL_WIDTH)


def generate_excel(filename: str, sheets: Any = None) -> dict:
    """Construit un classeur `.xlsx` et l'ÃĐcrit dans `DOWNLOADS_DIR`.

    Args:
        filename: nom souhaitÃĐ (assaini, extension forcÃĐe `.xlsx`).
        sheets: une feuille (dict) ou une liste de feuilles `{name, columns, rows}`.
            `rows` accepte une liste de listes OU une liste de dicts.

    Returns:
        `{"status": "ok", "filename", "path", "download_url", "sheets", "rows",
        "size"}` en cas de succÃĻs, sinon `{"error": "..."}`.
    """
    if not HAS_OPENPYXL:
        return {"error": "openpyxl non installÃĐ â `pip install openpyxl`."}

    norm = _normalize_sheets(sheets)
    if not norm:
        return {"error": "Aucune donnÃĐe : fournis au moins une feuille avec des lignes."}

    safe_name = _safe_filename(filename)
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest = (DOWNLOADS_DIR / safe_name).resolve()
    if dest.parent != DOWNLOADS_DIR.resolve():
        return {"error": f"Nom de fichier invalide : {filename!r}"}

    wb = Workbook()
    wb.remove(wb.active)  # on crÃĐe nos propres feuilles nommÃĐes

    used_titles: set[str] = set()
    sheet_names: list[str] = []
    total_rows = 0

    for sh in norm:
        title = _safe_sheet_title(sh["name"], used_titles)
        ws = wb.create_sheet(title=title)
        sheet_names.append(title)

        columns = sh.get("columns")
        rows = sh.get("rows") or []
        if len(rows) > _MAX_ROWS:
            logger.warning("Excel %s/%s : %d lignes tronquÃĐes Ã  %d",
                           safe_name, title, len(rows), _MAX_ROWS)
            rows = rows[:_MAX_ROWS]

        if columns:
            ws.append([_coerce(c) for c in list(columns)[:_MAX_COLS]])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"

        for r in rows:
            cells = r if isinstance(r, (list, tuple)) else [r]
            ws.append([_coerce(c) for c in list(cells)[:_MAX_COLS]])
            total_rows += 1

        _autosize(ws, columns, rows)

    if not wb.sheetnames:  # garde-fou : un classeur sans feuille ne peut ÃŠtre sauvÃĐ
        wb.create_sheet(title="Feuille1")
        sheet_names.append("Feuille1")

    wb.save(dest)
    size = dest.stat().st_size
    logger.info("Excel gÃĐnÃĐrÃĐ : %s (%d feuille(s), %d lignes, %d o)",
                safe_name, len(sheet_names), total_rows, size)

    return {
        "status": "ok",
        "filename": safe_name,
        "path": str(dest),
        "download_url": f"/api/files/{safe_name}",
        "sheets": sheet_names,
        "rows": total_rows,
        "size": size,
    }
